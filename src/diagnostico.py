import os
import time
import pandas as pd
from tkinter import messagebox
from graficos import *
from reporte import (
    inicializar_reporte,
    agregar_etapa_general,
    agregar_etapa_limpieza,
    agregar_etapa_graficos
)
from actualizar_bdm import actualizar_bdm
from pyspark.sql import SparkSession
from cargar_datos_spark import cargar_y_limpiar_datos_spark
from normalizacion_semantica import aplicar_normalizacion_semantica
from sugerencia_codigo_unspsc import sugerir_codigos_para_faltantes
from sospechosos_codigo_unspsc import detectar_sospechosos_unspsc
from agrupamiento_similar import agrupar_registros_similares
from evaluacion_calidad import evaluar_calidad_por_registro
from utils.registro_historial import registrar_diagnostico_bdm

spark = SparkSession.builder \
    .appName("DiagnosticoCatalogo") \
    .getOrCreate()

def procesar_excel(excel_path, ruta_guardado, barra_progreso, ventana):
    inicio_total = time.time()

    try:
        nombre_archivo = os.path.basename(excel_path)
        inicializar_reporte(nombre_archivo, ruta_guardado)

        # Limpieza y carga de datos
        inicio_limpieza = time.time()
        df_spark = cargar_y_limpiar_datos_spark(spark, excel_path)
        df = df_spark.toPandas()
        info_limpieza = {"metodo": "spark"}
        duracion_limpieza = time.time() - inicio_limpieza

        # Crear columnas temporales con términos ordenados alfabéticamente
        df["_desc_larga_ordenada"] = df["Descripcion Larga"].astype(str) \
            .apply(lambda x: ",".join(sorted(x.split(";"))))
        df["_desc_corta_ordenada"] = df["Descripcion Corta"].astype(str) \
            .apply(lambda x: ",".join(sorted(x.split(";"))))

        # Detectar duplicados
        dup_larga = df["_desc_larga_ordenada"].duplicated(keep=False)
        dup_corta = df["_desc_corta_ordenada"].duplicated(keep=False)

        # Contar duplicados
        dup_larga_count = dup_larga.sum()
        dup_corta_count = dup_corta.sum()

        # Identificar únicos (sin duplicados)
        val_unicos_larga = df[~dup_larga]["_desc_larga_ordenada"].nunique()
        val_unicos_corta = df[~dup_corta]["_desc_corta_ordenada"].nunique()

        # Ajustar estimación final
        val_unicos_larga_fin = val_unicos_larga + (dup_larga_count / 2)
        val_unicos_corta_fin = val_unicos_corta + (dup_corta_count / 2)

        # Generación de gráficos
        if barra_progreso:
            barra_progreso.set(0.7)
            ventana.update_idletasks()

        inicio_graficos = time.time()
        graficos_ok = []
        graficos_fallidos = []
        avance_por_grafico = (0.95 - 0.7) / 12

        def intentar(nombre, funcion):
            try:
                funcion(df, ruta_guardado)
                graficos_ok.append(nombre)
                if barra_progreso:
                    prog = barra_progreso.get() + avance_por_grafico
                    ventana.after(0, barra_progreso.set, min(prog, 0.95))
            except Exception:
                graficos_fallidos.append(nombre)

        # Listado de gráficos
        intentar("1. Porcentaje datos", graficar_porcentaje_datos)
        intentar("2. Descripciones únicas", graficar_descripciones_unicas)
        intentar("3. Descripciones duplicadas", graficar_duplicados_descripciones)
        intentar("4. Falta información", graficar_falta_info_descripcion_corta)
        intentar("5. Distribución de idioma", graficar_idiomas_descripcion_larga)
        intentar("6. Unidades medida", graficar_unidades_medida)
        intentar("7. Grupo UNSPSC", graficar_top_grupo_unspsc)
        intentar("8. Productos UNSPSC", graficar_top_producto_unspsc)
        intentar("9. Asignación producto", graficar_verificacion_unspsc)
        intentar("10. Cantidad de características", graficar_cantidad_caracteristicas)
        intentar("11. Código categoría", graficar_codigo_categoria)
        intentar("12. Tipo artículo", graficar_tipo_articulo)

        duracion_graficos = time.time() - inicio_graficos
        agregar_etapa_graficos(duracion_graficos, graficos_ok, graficos_fallidos)

        # Guardar DataFrame y duplicados en un Excel
        datos_dir = os.path.join(ruta_guardado, "datos")
        os.makedirs(datos_dir, exist_ok=True)
        ruta_excel = os.path.join(datos_dir, "datos_procesados.xlsx")

        duplicados_larga_df = df[dup_larga].copy()
        duplicados_corta_df = df[dup_corta].copy()
        duplicados_larga_df["Código Duplicado Larga"] = \
            duplicados_larga_df.groupby("_desc_larga_ordenada").ngroup() + 1
        duplicados_corta_df["Código Duplicado Corta"] = \
            duplicados_corta_df.groupby("_desc_corta_ordenada").ngroup() + 1

        # Crear o anexar Excel
        with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="w") as writer:
            df.head(1).to_excel(writer, sheet_name="Temp", index=False)
        with pd.ExcelWriter(ruta_excel, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name="Datos completos", index=False)
            duplicados_larga_df[["Código Duplicado Larga","Descripcion Larga"]] \
                .to_excel(writer, sheet_name="Duplicados Larga", index=False)
            duplicados_corta_df[["Código Duplicado Corta","Descripcion Corta"]] \
                .to_excel(writer, sheet_name="Duplicados Corta", index=False)

        # Eliminar hoja temporal
        from openpyxl import load_workbook
        try:
            wb = load_workbook(ruta_excel)
            if "Temp" in wb.sheetnames:
                wb.remove(wb["Temp"])
                wb.save(ruta_excel)
        except Exception:
            pass

        # Registrar exportación en el reporte
        with open(os.path.join(ruta_guardado, "reporte_diagnostico.txt"), "a", encoding="utf-8") as f:
            f.write(f"Archivo Excel exportado: {ruta_excel}\n\n")

        # Etapa general
        duracion_total = time.time() - inicio_total
        agregar_etapa_general(len(df), len(df.columns), duracion_total)

        # Normalización y valoración final
        df = aplicar_normalizacion_semantica(df)
        df = sugerir_codigos_para_faltantes(df)
        df = detectar_sospechosos_unspsc(df)
        df_similares = agrupar_registros_similares(df)
        report_dir = os.path.join("data","report")
        os.makedirs(report_dir, exist_ok=True)
        df_similares.to_excel(os.path.join(report_dir, "agrupamiento_similar.xlsx"), index=False)
        df = evaluar_calidad_por_registro(df)

        # — Crear carpeta processed y guardar catálogo final —
        carpeta_proc = os.path.join(os.getcwd(), "data", "processed")
        os.makedirs(carpeta_proc, exist_ok=True)
        ruta_cat = os.path.join(carpeta_proc, "catalogo_final.csv")
        df.to_csv(ruta_cat, index=False)

        # Actualizar BDM
        actualizar_bdm(
            ruta_catalogo=ruta_cat,
            ruta_codigos=os.path.join(carpeta_proc, "codigos_unspsc.csv"),
            ruta_mapeos=os.path.join(carpeta_proc, "mapeos_v14_v26.csv"),
            ruta_caracteristicas=os.path.join(carpeta_proc, "caracteristicas.csv")
        )

        # Mantener CSV raíz para compatibilidad
        df.to_csv("catalogo_final.csv", index=False)

        # Registrar historial en BDM de diagnóst**tico
        registrar_diagnostico_bdm(
            catalogo_nombre=nombre_archivo,
            total_rows=len(df),
            total_dups=(dup_larga_count + dup_corta_count),
            nuevos=df['codigo'].isna().sum(),
            actualizados=0,
            total_seg=duracion_total,
            estado="OK",
            notas="Ejecución finalizada correctamente"
        )

        messagebox.showinfo("Finalizado", "El diagnóstico ha finalizado correctamente.")

    except Exception as e:
        ventana.after(0, barra_progreso.set, 0)
        mensaje = str(e)
        ventana.after(0, lambda: messagebox.showerror("Error", mensaje))
